import streamlit as st
import pandas as pd
import io
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Order Assignment Automation",
    page_icon="📋",
    layout="wide"
)

st.markdown("""
<style>
    .main { background-color: #0f1117; }
    .block-container { padding-top: 1.5rem; max-width: 1200px; }
    h1 { color: #e2e8f0; font-size: 1.8rem; font-weight: 700; }
    h2, h3 { color: #cbd5e1; }
    .stButton > button {
        background: linear-gradient(135deg, #6366f1, #8b5cf6);
        color: white; border: none; border-radius: 8px;
        padding: 0.5rem 1.5rem; font-weight: 600;
        transition: all 0.2s;
    }
    .stButton > button:hover { opacity: 0.85; transform: translateY(-1px); }
    .metric-card {
        background: #1e2130; border: 1px solid #2d3148;
        border-radius: 10px; padding: 1rem 1.25rem; text-align: center;
    }
    .metric-card .val { font-size: 2rem; font-weight: 700; color: #818cf8; }
    .metric-card .lbl { font-size: 0.8rem; color: #94a3b8; margin-top: 2px; }
    .section-box {
        background: #1e2130; border: 1px solid #2d3148;
        border-radius: 12px; padding: 1.25rem 1.5rem; margin-bottom: 1rem;
    }
    .step-header {
        font-size: 0.85rem; font-weight: 700; color: #818cf8;
        text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 0.5rem;
    }
    .part-header {
        font-size: 1rem; font-weight: 700; color: #e2e8f0;
        background: #1e2130; border-left: 4px solid #6366f1;
        padding: 0.6rem 1rem; border-radius: 0 8px 8px 0; margin: 1rem 0 0.5rem 0;
    }
    div[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }
    .stTabs [data-baseweb="tab"] { color: #94a3b8; }
    .stTabs [aria-selected="true"] { color: #818cf8 !important; }
    .info-chip {
        background:#1e2130; border:1px solid #2d3148; border-radius:6px;
        padding:6px 12px; font-size:0.8rem; color:#94a3b8; display:inline-block; margin:3px;
    }
    .shift-box-a {
        background:#1a1f35; border:2px solid #6366f1; border-radius:10px;
        padding:12px 16px; margin-bottom:8px;
    }
    .shift-box-b {
        background:#1a2e22; border:2px solid #10b981; border-radius:10px;
        padding:12px 16px; margin-bottom:8px;
    }
    .shift-label-a { color:#818cf8; font-weight:700; font-size:0.95rem; }
    .shift-label-b { color:#34d399; font-weight:700; font-size:0.95rem; }
</style>
""", unsafe_allow_html=True)

# ─── Column definitions ───────────────────────────────────────────────────────
COMMON_COLS = ["Status", "Additional Notes", "Bundled Makegood?"]
FULL_CONTRACT_COLS = COMMON_COLS + [
    "Next MG Date - CIOC USE ONLY", "Assigned TF Coordinator - TF USE ONLY",
    "Start Date", "End Date", "Region", "TIM#", "Client Name",
    "Default Makegood Parameters", "Can PE$ Move to Next Month?",
    "Makegood Only on Zone Spots Dropped"
]
RATING_COLS = COMMON_COLS + [
    "Next MG Date - CIOC USE ONLY", "Assigned TF Coordinator - TF USE ONLY",
    "Start Date", "End Date", "Region", "TIM#", "Client Name",
    "Default Makegood Parameters", "Can PE$ Move to Next Month?",
    "MakeGood Hiatus Weeks"
]
AD_HOC_COLS = COMMON_COLS + [
    "Created On", "Modified On", "Modified By",
    "Next MG Date - CIOC USE ONLY", "Assigned TF Coordinator - TF USE ONLY",
    "Start Date", "End Date", "TIM#", "Client Name", "Default Makegood Parameters"
]

# ─── Session state ────────────────────────────────────────────────────────────
defaults = {
    "raw_df": None, "processed": None,
    "order_type": "Full Contract",
    "name_sets": [[]], "percentages": [100], "num_sets": 1,
    # Slot assignment
    "slot_raw_df": None, "slot_processed": None,
    "num_slots": 6, "slot_configs": [],
    # Shift split (part 2)
    "shift_a_pct": 50,
    "shift_a_name_sets": [[]], "shift_a_percentages": [100], "shift_a_num_sets": 1,
    "shift_b_name_sets": [[]], "shift_b_percentages": [100], "shift_b_num_sets": 1,
    # Political
    "pol_raw_df": None, "pol_processed": None, "pol_names": [],
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─── Shared helpers ───────────────────────────────────────────────────────────
def detect_order_type(df):
    cols = [c.strip() for c in df.columns]
    if "MakeGood Hiatus Weeks" in cols:
        return "Rating"
    elif "Created On" in cols or "Modified On" in cols:
        return "Ad Hoc"
    return "Full Contract"

def normalize_columns(df, order_type):
    ref = {"Full Contract": FULL_CONTRACT_COLS,
           "Rating": RATING_COLS,
           "Ad Hoc": AD_HOC_COLS}[order_type]
    col_map = {}
    for col in df.columns:
        for rc in ref:
            if col.strip().lower() == rc.strip().lower():
                col_map[col] = rc
    return df.rename(columns=col_map)

def assign_round_robin(df, name_sets, percentages):
    all_names = [n for ns in name_sets for n in ns]
    if not all_names:
        return df
    df = df.copy()
    df["Names"] = ""
    total = len(df)
    idx = 0
    for i, (names, pct) in enumerate(zip(name_sets, percentages)):
        if not names:
            continue
        count = (total - idx) if i == len(name_sets) - 1 else round(total * pct / 100)
        count = min(count, total - idx)
        slice_indices = list(df.index[idx: idx + count])
        n = len(names)
        for pos, orig_idx in enumerate(slice_indices):
            df.at[orig_idx, "Names"] = names[pos % n]
        idx += count
    return df

def get_col(df, col_name):
    if col_name in df.columns:
        return df[col_name]
    return pd.Series([""] * len(df), index=df.index)

def build_output(df, order_type):
    today = date.today()
    out = pd.DataFrame(index=df.index)
    out["Current Date"]                     = str(today)
    out["Current Month"]                    = today.strftime("%B %Y")
    out["Names"]                            = get_col(df, "Names")
    out["Region"]                           = get_col(df, "Region")
    out["TIM#"]                             = get_col(df, "TIM#")
    out["CM Name"]                          = get_col(df, "Requestor Email")
    out["Status"]                           = get_col(df, "Status")
    out["Time Zone of Order"]               = get_col(df, "Time Zone of Order")
    out["Ad Hoc / Ad Hoc or Full Contract"] = order_type
    return out.reset_index(drop=True)

def style_excel(ws):
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="B0B0B0")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even    = PatternFill("solid", fgColor="DCE6F1")
    fill_odd     = PatternFill("solid", fgColor="FFFFFF")
    max_col, max_row = ws.max_column, ws.max_row
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center_align; cell.border = border
    for r in range(2, max_row + 1):
        fill = fill_even if r % 2 == 0 else fill_odd
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill; cell.alignment = left_align; cell.border = border
    for c in range(1, max_col + 1):
        ltr = get_column_letter(c)
        max_len = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, max_row + 1)),
            default=8
        )
        ws.column_dimensions[ltr].width = min(max_len + 4, 42)
    ws.row_dimensions[1].height = 30

def style_excel_green(ws):
    """Green header variant for 7-4 shift sheets."""
    header_fill  = PatternFill("solid", fgColor="1A4731")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="B0B0B0")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even    = PatternFill("solid", fgColor="D6F0E0")
    fill_odd     = PatternFill("solid", fgColor="FFFFFF")
    max_col, max_row = ws.max_column, ws.max_row
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center_align; cell.border = border
    for r in range(2, max_row + 1):
        fill = fill_even if r % 2 == 0 else fill_odd
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill; cell.alignment = left_align; cell.border = border
    for c in range(1, max_col + 1):
        ltr = get_column_letter(c)
        max_len = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, max_row + 1)),
            default=8
        )
        ws.column_dimensions[ltr].width = min(max_len + 4, 42)
    ws.row_dimensions[1].height = 30

def make_excel_bytes(df_out, pivot_df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Assigned Orders")
        if pivot_df is not None and not pivot_df.empty:
            pivot_df.to_excel(writer, index=False, sheet_name="TIM Pivot")
    buf.seek(0)
    wb = load_workbook(buf)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_row > 0 and ws.max_column > 0:
            style_excel(ws)
    out_buf = io.BytesIO()
    wb.save(out_buf); out_buf.seek(0)
    return out_buf

def make_slot_excel_bytes(slot_sheets, pivot_df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pivot_df.to_excel(writer, index=False, sheet_name="Pivot - All Slots")
        for slot_num, df_slot in enumerate(slot_sheets, start=1):
            df_slot.to_excel(writer, index=False, sheet_name=f"Slot {slot_num}")
    buf.seek(0)
    wb = load_workbook(buf)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_row > 0 and ws.max_column > 0:
            style_excel(ws)
    out_buf = io.BytesIO()
    wb.save(out_buf); out_buf.seek(0)
    return out_buf

def make_shift_excel_bytes(
    df_a, df_b,
    pivot_a, pivot_b,
    slot_sheets_a, slot_sheets_b,
    num_slots_a, num_slots_b
):
    """
    Sheet layout:
      1. Pivot 5-2  (blue header)
      2. Pivot 7-4  (green header)
      3. 5-2 Slot 1, 5-2 Slot 2 … (blue)
      4. 7-4 Slot 1, 7-4 Slot 2 … (green)
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Pivot sheets
        pivot_a.to_excel(writer, index=False, sheet_name="Pivot 5-2")
        pivot_b.to_excel(writer, index=False, sheet_name="Pivot 7-4")
        # 5-2 slot sheets
        for i, df_slot in enumerate(slot_sheets_a, start=1):
            df_slot.to_excel(writer, index=False, sheet_name=f"5-2 Slot {i}")
        # 7-4 slot sheets
        for i, df_slot in enumerate(slot_sheets_b, start=1):
            df_slot.to_excel(writer, index=False, sheet_name=f"7-4 Slot {i}")

    buf.seek(0)
    wb = load_workbook(buf)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.max_row > 0 and ws.max_column > 0:
            if "7-4" in sheet:
                style_excel_green(ws)
            else:
                style_excel(ws)
    out_buf = io.BytesIO()
    wb.save(out_buf); out_buf.seek(0)
    return out_buf


def render_shift_name_sets(shift_key, shift_label, color):
    """Render name-set inputs for one shift. Returns (name_sets, percentages, pct_total)."""
    num_key  = f"{shift_key}_num_sets"
    ns_key   = f"{shift_key}_name_sets"
    pct_key  = f"{shift_key}_percentages"

    ns_val = st.number_input(
        f"Name sets for {shift_label}", min_value=1, max_value=5,
        value=st.session_state[num_key], key=f"{shift_key}_ns_input"
    )
    st.session_state[num_key] = int(ns_val)

    while len(st.session_state[ns_key]) < ns_val:
        st.session_state[ns_key].append([])
        st.session_state[pct_key].append(0)
    st.session_state[ns_key]  = st.session_state[ns_key][:ns_val]
    st.session_state[pct_key] = st.session_state[pct_key][:ns_val]

    cols = st.columns(ns_val)
    cur_names = []; cur_pcts = []; pct_total = 0
    for j, col in enumerate(cols):
        with col:
            st.markdown(
                f'<div style="background:{color}22;border:1px solid {color}44;border-radius:6px;'
                f'padding:6px 10px;margin-bottom:4px;">'
                f'<span style="color:{color};font-weight:700;font-size:0.8rem;">Set {j+1}</span></div>',
                unsafe_allow_html=True
            )
            pct = st.number_input("% of shift orders", min_value=0, max_value=100,
                                  value=st.session_state[pct_key][j],
                                  key=f"{shift_key}_pct_{j}")
            cur_pcts.append(int(pct)); pct_total += pct
            names_raw = st.text_area(
                "Names (one per line)",
                value="\n".join(st.session_state[ns_key][j]),
                height=120, key=f"{shift_key}_names_{j}",
                placeholder="Alice\nBob\nCarol"
            )
            cur_names.append([n.strip() for n in names_raw.splitlines() if n.strip()])

    st.session_state[ns_key]  = cur_names
    st.session_state[pct_key] = cur_pcts
    return cur_names, cur_pcts, pct_total


# ═══════════════════════════════════════════════════════════════════════════════
#  UI
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("## 📋 Order Assignment Automation")
st.divider()

part1_tab, part2_tab, part3_tab = st.tabs([
    "📂 Part 1 — Standard Assignment",
    "🗂️ Part 2 — Slot-wise Assignment",
    "🏛️ Part 3 — Political Orders",
])


# ══════════════════════════════════════════════════════════════════════════════
#  PART 1 — unchanged
# ══════════════════════════════════════════════════════════════════════════════
with part1_tab:
    st.markdown('<div class="step-header">Step 1 — Upload assignment sheet</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Drop your Excel or CSV file here", type=["xlsx","xls","csv"],
                                label_visibility="collapsed", key="p1_upload")
    if uploaded is not None:
        try:
            df_raw = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
            st.session_state.raw_df     = df_raw
            st.session_state.order_type = detect_order_type(df_raw)
            st.session_state.processed  = None
            st.success(f"✅ Loaded **{len(df_raw):,}** rows")
        except Exception as e:
            st.error(f"Could not read file: {e}")

    if st.session_state.raw_df is not None:
        df = st.session_state.raw_df.copy()
        st.divider()
        st.markdown('<div class="step-header">Step 2 — Configure order type & name lists</div>', unsafe_allow_html=True)
        col_ot, _ = st.columns([2, 5])
        with col_ot:
            order_type = st.selectbox("Order type", ["Full Contract","Rating","Ad Hoc"],
                index=["Full Contract","Rating","Ad Hoc"].index(st.session_state.order_type), key="p1_otype")
            st.session_state.order_type = order_type

        st.markdown("**Name sets & percentage splits**")
        st.caption("Orders are distributed round-robin (max 1 order difference per person). Percentages must total 100.")
        num_sets = st.number_input("Number of name sets", min_value=1, max_value=5,
                                   value=st.session_state.num_sets, key="num_sets_input")
        st.session_state.num_sets = int(num_sets)
        while len(st.session_state.name_sets) < num_sets:
            st.session_state.name_sets.append([]); st.session_state.percentages.append(0)
        st.session_state.name_sets   = st.session_state.name_sets[:num_sets]
        st.session_state.percentages = st.session_state.percentages[:num_sets]

        badge_colors = ["#6366f1","#10b981","#f59e0b","#ef4444","#8b5cf6"]
        grid_cols = st.columns(num_sets)
        pct_total = 0; current_names = []; current_pcts = []
        for i, col in enumerate(grid_cols):
            with col:
                bc = badge_colors[i % 5]
                st.markdown(f'<div style="background:{bc}22;border:1px solid {bc}44;border-radius:8px;padding:8px 12px;margin-bottom:6px;"><span style="color:{bc};font-weight:700;font-size:0.85rem;">Set {i+1}</span></div>', unsafe_allow_html=True)
                pct = st.number_input("% of orders", min_value=0, max_value=100,
                                      value=st.session_state.percentages[i], key=f"pct_{i}")
                current_pcts.append(int(pct)); pct_total += pct
                names_raw = st.text_area("Names (one per line)", value="\n".join(st.session_state.name_sets[i]),
                                         height=130, key=f"names_{i}", placeholder="Alice\nBob\nCarol")
                current_names.append([n.strip() for n in names_raw.splitlines() if n.strip()])
        st.session_state.name_sets   = current_names
        st.session_state.percentages = current_pcts

        if pct_total != 100:
            st.warning(f"⚠️ Percentages sum to **{pct_total}%** — must total 100%.")
        else:
            st.success("✅ Percentages sum to 100%")
        st.divider()

        st.markdown('<div class="step-header">Step 3 — Process & preview</div>', unsafe_allow_html=True)
        all_names_flat = [n for ns in st.session_state.name_sets for n in ns]
        can_run = len(all_names_flat) > 0 and pct_total == 100
        if not can_run:
            st.info("ℹ️ Enter names and make sure percentages total 100% to enable Run Assignment.")

        if st.button("🚀 Run Assignment", disabled=not can_run, key="p1_run"):
            try:
                df_norm     = normalize_columns(df, st.session_state.order_type)
                df_assigned = assign_round_robin(df_norm, st.session_state.name_sets, st.session_state.percentages)
                df_out      = build_output(df_assigned, st.session_state.order_type)
                st.session_state.processed = df_out
                st.success(f"✅ Assignment complete — {len(df_out):,} rows processed.")
            except Exception as e:
                st.error(f"Assignment failed: {e}"); st.exception(e)

        if st.session_state.processed is not None:
            df_out     = st.session_state.processed
            total_rows = len(df_out)
            assigned   = int((df_out["Names"] != "").sum())
            unassigned = total_rows - assigned
            tim_count  = int(df_out["TIM#"].nunique()) if "TIM#" in df_out.columns else 0
            st.markdown("")
            m1,m2,m3,m4 = st.columns(4)
            for col,val,lbl in [(m1,total_rows,"Total Orders"),(m2,assigned,"Assigned"),(m3,unassigned,"Unassigned"),(m4,tim_count,"Unique TIM#")]:
                with col:
                    st.markdown(f'<div class="metric-card"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>', unsafe_allow_html=True)
            st.markdown("")
            pivot_df = pd.DataFrame()
            if "Names" in df_out.columns and "TIM#" in df_out.columns:
                pivot_df = (df_out[df_out["Names"]!=""].groupby("Names")["TIM#"].count()
                            .reset_index().rename(columns={"TIM#":"Total TIM# Count"})
                            .sort_values("Total TIM# Count",ascending=False).reset_index(drop=True))
            tab1,tab2,tab3 = st.tabs(["📄 Assigned Output","📊 TIM# Pivot","🔍 Raw Data"])
            with tab1: st.dataframe(df_out, use_container_width=True, height=400)
            with tab2:
                if not pivot_df.empty: st.dataframe(pivot_df, use_container_width=True, height=400)
                else: st.info("No pivot data available.")
            with tab3: st.dataframe(st.session_state.raw_df, use_container_width=True, height=400)
            st.divider()
            st.markdown('<div class="step-header">Step 4 — Export</div>', unsafe_allow_html=True)
            ec1,ec2 = st.columns(2)
            with ec1:
                try:
                    xlsx_bytes = make_excel_bytes(df_out, pivot_df.copy() if not pivot_df.empty else pd.DataFrame())
                    st.download_button("⬇️ Download Excel (.xlsx)", data=xlsx_bytes,
                                       file_name=f"assigned_orders_{date.today()}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True, key="p1_dl_xlsx")
                except Exception as e:
                    st.error(f"Excel generation failed: {e}"); st.exception(e)
            with ec2:
                st.download_button("⬇️ Download CSV", data=df_out.to_csv(index=False).encode("utf-8"),
                                   file_name=f"assigned_orders_{date.today()}.csv", mime="text/csv",
                                   use_container_width=True, key="p1_dl_csv")
    else:
        st.markdown("""
        <div class="section-box" style="text-align:center;padding:2.5rem;">
            <div style="font-size:2.5rem;margin-bottom:0.5rem;">📂</div>
            <div style="color:#94a3b8;font-size:0.95rem;">
                Upload an <strong style="color:#cbd5e1;">Excel or CSV</strong> file above to get started.
            </div>
            <hr style="border-color:#2d3148;margin:1.5rem auto;width:60%;">
            <div style="display:flex;justify-content:center;flex-wrap:wrap;gap:6px;">
                <span class="info-chip">✔ Auto-detects order type</span>
                <span class="info-chip">✔ True even round-robin distribution</span>
                <span class="info-chip">✔ Up to 5 name-set splits</span>
                <span class="info-chip">✔ Styled Excel with bold headers & borders</span>
                <span class="info-chip">✔ TIM# pivot table</span>
            </div>
        </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PART 2 — Slot-wise Assignment with Shift Split
# ══════════════════════════════════════════════════════════════════════════════
with part2_tab:
    st.markdown('<div class="step-header">Step 1 — Upload order sheet</div>', unsafe_allow_html=True)
    slot_uploaded = st.file_uploader("Drop your Excel or CSV file here", type=["xlsx","xls","csv"],
                                     label_visibility="collapsed", key="p2_upload")
    if slot_uploaded is not None:
        try:
            slot_raw = pd.read_csv(slot_uploaded) if slot_uploaded.name.endswith(".csv") else pd.read_excel(slot_uploaded)
            st.session_state.slot_raw_df    = slot_raw
            st.session_state.slot_processed = None
            st.success(f"✅ Loaded **{len(slot_raw):,}** rows")
        except Exception as e:
            st.error(f"Could not read file: {e}")

    if st.session_state.slot_raw_df is not None:
        slot_df      = st.session_state.slot_raw_df.copy()
        total_orders = len(slot_df)

        st.divider()
        st.markdown('<div class="step-header">Step 2 — Shift split & slot configuration</div>', unsafe_allow_html=True)

        col_ot2, _ = st.columns([2, 5])
        with col_ot2:
            slot_order_type = st.selectbox("Order type", ["Full Contract","Rating","Ad Hoc"], key="p2_otype")

        # ── Shift percentage split ────────────────────────────────────────────
        st.markdown("#### 🕐 Shift Order Split")
        spl_col1, spl_col2, spl_col3 = st.columns([2, 2, 3])
        with spl_col1:
            shift_a_pct = st.number_input("5-2 shift  (% of total orders)",
                                          min_value=0, max_value=100,
                                          value=st.session_state.shift_a_pct, key="shift_a_pct_input")
            st.session_state.shift_a_pct = int(shift_a_pct)
        shift_b_pct = 100 - shift_a_pct
        with spl_col2:
            st.metric("7-4 shift  (% of total orders)", f"{shift_b_pct}%")

        orders_a = round(total_orders * shift_a_pct / 100)
        orders_b = total_orders - orders_a
        with spl_col3:
            st.info(f"🔵 **5-2**: {orders_a:,} orders  |  🟢 **7-4**: {orders_b:,} orders")

        st.divider()

        # ── Slot counts per shift ─────────────────────────────────────────────
        sc1, sc2 = st.columns(2)
        with sc1:
            num_slots_a = st.number_input("Number of slots for 5-2 shift",
                                          min_value=1, max_value=20, value=6, key="p2_slots_a")
        with sc2:
            num_slots_b = st.number_input("Number of slots for 7-4 shift",
                                          min_value=1, max_value=20, value=6, key="p2_slots_b")

        num_slots_a = int(num_slots_a)
        num_slots_b = int(num_slots_b)

        ops_a = orders_a // num_slots_a; rem_a = orders_a % num_slots_a
        ops_b = orders_b // num_slots_b; rem_b = orders_b % num_slots_b

        ic1, ic2 = st.columns(2)
        with ic1:
            st.info(f"🔵 5-2: {orders_a} orders ÷ {num_slots_a} slots = {ops_a}/slot"
                    + (f" (first {rem_a} get +1)" if rem_a else " (even)"))
        with ic2:
            st.info(f"🟢 7-4: {orders_b} orders ÷ {num_slots_b} slots = {ops_b}/slot"
                    + (f" (first {rem_b} get +1)" if rem_b else " (even)"))

        # ── Ensure slot_configs lists long enough ─────────────────────────────
        total_slots_needed = num_slots_a + num_slots_b
        while len(st.session_state.slot_configs) < total_slots_needed:
            st.session_state.slot_configs.append(
                {"num_sets": 1, "name_sets": [[]], "percentages": [100]}
            )
        # We'll index 0..num_slots_a-1 for 5-2, num_slots_a..total-1 for 7-4

        st.divider()
        all_slots_valid = True

        # ── 5-2 Shift slot configuration ─────────────────────────────────────
        st.markdown('<div class="shift-box-a"><span class="shift-label-a">🔵 5-2 Shift — Configure Slots</span></div>', unsafe_allow_html=True)

        for s in range(num_slots_a):
            cfg = st.session_state.slot_configs[s]
            slot_size = ops_a + (1 if s < rem_a else 0)
            with st.expander(f"🔵 5-2 · Slot {s+1}  —  {slot_size} orders", expanded=(s == 0)):
                sn = st.number_input(f"Name sets in 5-2 Slot {s+1}", min_value=1, max_value=5,
                                     value=cfg.get("num_sets", 1), key=f"sa{s}_num_sets")
                cfg["num_sets"] = int(sn)
                while len(cfg["name_sets"]) < sn:
                    cfg["name_sets"].append([]); cfg["percentages"].append(0)
                cfg["name_sets"]   = cfg["name_sets"][:sn]
                cfg["percentages"] = cfg["percentages"][:sn]

                s_cols = st.columns(sn)
                s_pct_total = 0; s_cur_names = []; s_cur_pcts = []
                for j, sc in enumerate(s_cols):
                    with sc:
                        st.markdown(f'<div style="background:#6366f122;border:1px solid #6366f144;border-radius:6px;padding:6px 10px;margin-bottom:4px;"><span style="color:#818cf8;font-weight:700;font-size:0.8rem;">Set {j+1}</span></div>', unsafe_allow_html=True)
                        sp = st.number_input("% of slot orders", min_value=0, max_value=100,
                                             value=cfg["percentages"][j], key=f"sa{s}_pct_{j}")
                        s_cur_pcts.append(int(sp)); s_pct_total += sp
                        sn_raw = st.text_area("Names (one per line)",
                                              value="\n".join(cfg["name_sets"][j]),
                                              height=110, key=f"sa{s}_names_{j}", placeholder="Alice\nBob")
                        s_cur_names.append([n.strip() for n in sn_raw.splitlines() if n.strip()])
                cfg["name_sets"]   = s_cur_names
                cfg["percentages"] = s_cur_pcts

                if s_pct_total != 100:
                    st.warning(f"⚠️ 5-2 Slot {s+1} percentages = {s_pct_total}% — must be 100%.")
                    all_slots_valid = False
                else:
                    flat = [n for ns in s_cur_names for n in ns]
                    if flat:
                        st.success(f"✅ 5-2 Slot {s+1} ready — {len(flat)} people, {slot_size} orders")
                    else:
                        st.warning(f"⚠️ 5-2 Slot {s+1} has no names yet."); all_slots_valid = False

            st.session_state.slot_configs[s] = cfg

        st.divider()

        # ── 7-4 Shift slot configuration ─────────────────────────────────────
        st.markdown('<div class="shift-box-b"><span class="shift-label-b">🟢 7-4 Shift — Configure Slots</span></div>', unsafe_allow_html=True)

        for s in range(num_slots_b):
            cfg_idx = num_slots_a + s
            cfg = st.session_state.slot_configs[cfg_idx]
            slot_size = ops_b + (1 if s < rem_b else 0)
            with st.expander(f"🟢 7-4 · Slot {s+1}  —  {slot_size} orders", expanded=(s == 0)):
                sn = st.number_input(f"Name sets in 7-4 Slot {s+1}", min_value=1, max_value=5,
                                     value=cfg.get("num_sets", 1), key=f"sb{s}_num_sets")
                cfg["num_sets"] = int(sn)
                while len(cfg["name_sets"]) < sn:
                    cfg["name_sets"].append([]); cfg["percentages"].append(0)
                cfg["name_sets"]   = cfg["name_sets"][:sn]
                cfg["percentages"] = cfg["percentages"][:sn]

                s_cols = st.columns(sn)
                s_pct_total = 0; s_cur_names = []; s_cur_pcts = []
                for j, sc in enumerate(s_cols):
                    with sc:
                        st.markdown(f'<div style="background:#10b98122;border:1px solid #10b98144;border-radius:6px;padding:6px 10px;margin-bottom:4px;"><span style="color:#34d399;font-weight:700;font-size:0.8rem;">Set {j+1}</span></div>', unsafe_allow_html=True)
                        sp = st.number_input("% of slot orders", min_value=0, max_value=100,
                                             value=cfg["percentages"][j], key=f"sb{s}_pct_{j}")
                        s_cur_pcts.append(int(sp)); s_pct_total += sp
                        sn_raw = st.text_area("Names (one per line)",
                                              value="\n".join(cfg["name_sets"][j]),
                                              height=110, key=f"sb{s}_names_{j}", placeholder="Alice\nBob")
                        s_cur_names.append([n.strip() for n in sn_raw.splitlines() if n.strip()])
                cfg["name_sets"]   = s_cur_names
                cfg["percentages"] = s_cur_pcts

                if s_pct_total != 100:
                    st.warning(f"⚠️ 7-4 Slot {s+1} percentages = {s_pct_total}% — must be 100%.")
                    all_slots_valid = False
                else:
                    flat = [n for ns in s_cur_names for n in ns]
                    if flat:
                        st.success(f"✅ 7-4 Slot {s+1} ready — {len(flat)} people, {slot_size} orders")
                    else:
                        st.warning(f"⚠️ 7-4 Slot {s+1} has no names yet."); all_slots_valid = False

            st.session_state.slot_configs[cfg_idx] = cfg

        st.divider()

        # ── Run ───────────────────────────────────────────────────────────────
        st.markdown('<div class="step-header">Step 3 — Run slot assignment</div>', unsafe_allow_html=True)
        if not all_slots_valid:
            st.info("ℹ️ Fix warnings above to enable Run Slot Assignment.")

        if st.button("🚀 Run Slot Assignment", disabled=not all_slots_valid, key="p2_run"):
            try:
                norm_df = normalize_columns(slot_df, slot_order_type)
                norm_df = norm_df.reset_index(drop=True)

                chunk_a = norm_df.iloc[:orders_a].copy().reset_index(drop=True)
                chunk_b = norm_df.iloc[orders_a:].copy().reset_index(drop=True)

                # Process 5-2 slots
                slot_results_a = []
                idx = 0
                for s in range(num_slots_a):
                    cfg = st.session_state.slot_configs[s]
                    sz  = ops_a + (1 if s < rem_a else 0)
                    chunk = chunk_a.iloc[idx: idx + sz].copy().reset_index(drop=True)
                    assigned = assign_round_robin(chunk, cfg["name_sets"], cfg["percentages"])
                    out = build_output(assigned, slot_order_type)
                    out.insert(0, "Slot", s + 1)
                    out.insert(1, "Shift", "5-2")
                    slot_results_a.append(out)
                    idx += sz

                # Process 7-4 slots
                slot_results_b = []
                idx = 0
                for s in range(num_slots_b):
                    cfg = st.session_state.slot_configs[num_slots_a + s]
                    sz  = ops_b + (1 if s < rem_b else 0)
                    chunk = chunk_b.iloc[idx: idx + sz].copy().reset_index(drop=True)
                    assigned = assign_round_robin(chunk, cfg["name_sets"], cfg["percentages"])
                    out = build_output(assigned, slot_order_type)
                    out.insert(0, "Slot", s + 1)
                    out.insert(1, "Shift", "7-4")
                    slot_results_b.append(out)
                    idx += sz

                all_a = pd.concat(slot_results_a, ignore_index=True) if slot_results_a else pd.DataFrame()
                all_b = pd.concat(slot_results_b, ignore_index=True) if slot_results_b else pd.DataFrame()

                def make_pivot(df):
                    if df.empty or "Names" not in df.columns:
                        return pd.DataFrame()
                    return (df[df["Names"] != ""].groupby("Names")["TIM#"].count()
                            .reset_index().rename(columns={"TIM#": "Total Order Count"})
                            .sort_values("Total Order Count", ascending=False)
                            .reset_index(drop=True))

                st.session_state.slot_processed = {
                    "slots_a": slot_results_a, "slots_b": slot_results_b,
                    "all_a": all_a, "all_b": all_b,
                    "pivot_a": make_pivot(all_a), "pivot_b": make_pivot(all_b),
                    "num_slots_a": num_slots_a, "num_slots_b": num_slots_b,
                }
                st.success(f"✅ Done — 🔵 5-2: {len(all_a):,} orders  |  🟢 7-4: {len(all_b):,} orders")
            except Exception as e:
                st.error(f"Slot assignment failed: {e}"); st.exception(e)

        # ── Results ───────────────────────────────────────────────────────────
        if st.session_state.slot_processed is not None:
            res = st.session_state.slot_processed
            slots_a  = res["slots_a"]; slots_b  = res["slots_b"]
            all_a    = res["all_a"];   all_b    = res["all_b"]
            pivot_a  = res["pivot_a"]; pivot_b  = res["pivot_b"]
            nsa      = res["num_slots_a"]; nsb = res["num_slots_b"]

            st.markdown("")
            mc1,mc2,mc3,mc4 = st.columns(4)
            with mc1: st.markdown(f'<div class="metric-card"><div class="val">{len(all_a)+len(all_b):,}</div><div class="lbl">Total Orders</div></div>', unsafe_allow_html=True)
            with mc2: st.markdown(f'<div class="metric-card"><div class="val" style="color:#818cf8">{len(all_a):,}</div><div class="lbl">🔵 5-2 Orders</div></div>', unsafe_allow_html=True)
            with mc3: st.markdown(f'<div class="metric-card"><div class="val" style="color:#34d399">{len(all_b):,}</div><div class="lbl">🟢 7-4 Orders</div></div>', unsafe_allow_html=True)
            with mc4: st.markdown(f'<div class="metric-card"><div class="val">{nsa+nsb}</div><div class="lbl">Total Slots</div></div>', unsafe_allow_html=True)
            st.markdown("")

            # Tab layout: Pivot 5-2 | Pivot 7-4 | 5-2 Slot1..N | 7-4 Slot1..N
            tab_labels = (
                ["📊 Pivot 5-2", "📊 Pivot 7-4"]
                + [f"🔵 5-2 Slot {i+1}" for i in range(len(slots_a))]
                + [f"🟢 7-4 Slot {i+1}" for i in range(len(slots_b))]
            )
            tabs = st.tabs(tab_labels)

            with tabs[0]:
                st.caption("5-2 Shift — order count per person")
                st.dataframe(pivot_a, use_container_width=True, height=380)
            with tabs[1]:
                st.caption("7-4 Shift — order count per person")
                st.dataframe(pivot_b, use_container_width=True, height=380)
            for i, (tab, df_slot) in enumerate(zip(tabs[2: 2+len(slots_a)], slots_a)):
                with tab:
                    st.caption(f"5-2 Shift · Slot {i+1} — {len(df_slot)} orders")
                    st.dataframe(df_slot, use_container_width=True, height=370)
            for i, (tab, df_slot) in enumerate(zip(tabs[2+len(slots_a):], slots_b)):
                with tab:
                    st.caption(f"7-4 Shift · Slot {i+1} — {len(df_slot)} orders")
                    st.dataframe(df_slot, use_container_width=True, height=370)

            st.divider()
            st.markdown('<div class="step-header">Step 4 — Export</div>', unsafe_allow_html=True)
            try:
                shift_xlsx = make_shift_excel_bytes(
                    all_a, all_b, pivot_a, pivot_b,
                    slots_a, slots_b, nsa, nsb
                )
                st.download_button(
                    "⬇️ Download Full Shift Excel (.xlsx)",
                    data=shift_xlsx,
                    file_name=f"shift_slot_assignment_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="p2_dl_xlsx"
                )
                st.caption("Excel contains: Pivot 5-2 · Pivot 7-4 · 5-2 Slot 1..N · 7-4 Slot 1..N")
            except Exception as e:
                st.error(f"Excel generation failed: {e}"); st.exception(e)

    else:
        st.markdown("""
        <div class="section-box" style="text-align:center;padding:2rem;">
            <div style="font-size:2rem;margin-bottom:0.4rem;">🗂️</div>
            <div style="color:#94a3b8;font-size:0.9rem;">
                Upload your order sheet above.<br>
                Set the 5-2 / 7-4 split %, number of slots per shift, configure names, then run.
            </div>
            <hr style="border-color:#2d3148;margin:1rem auto;width:50%;">
            <div style="display:flex;justify-content:center;flex-wrap:wrap;gap:6px;">
                <span class="info-chip">✔ Split orders between 5-2 and 7-4 shifts</span>
                <span class="info-chip">✔ Individual slots per shift</span>
                <span class="info-chip">✔ Separate pivot tables per shift</span>
                <span class="info-chip">✔ One Excel sheet per slot</span>
            </div>
        </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PART 3 — Political Orders (unchanged)
# ══════════════════════════════════════════════════════════════════════════════
with part3_tab:
    st.markdown('<div class="step-header">Step 1 — Upload political orders sheet</div>', unsafe_allow_html=True)
    pol_uploaded = st.file_uploader("Drop your Excel or CSV file here", type=["xlsx","xls","csv"],
                                    label_visibility="collapsed", key="p3_upload")
    if pol_uploaded is not None:
        try:
            pol_raw = pd.read_csv(pol_uploaded) if pol_uploaded.name.endswith(".csv") else pd.read_excel(pol_uploaded)
            st.session_state.pol_raw_df    = pol_raw
            st.session_state.pol_processed = None
            st.success(f"✅ Loaded **{len(pol_raw):,}** political orders")
        except Exception as e:
            st.error(f"Could not read file: {e}")

    if st.session_state.pol_raw_df is not None:
        pol_df = st.session_state.pol_raw_df.copy()
        st.divider()
        st.markdown('<div class="step-header">Step 2 — Configure</div>', unsafe_allow_html=True)
        col_pol1, _ = st.columns([2, 5])
        with col_pol1:
            pol_order_type = st.selectbox("Order type", ["Full Contract","Rating","Ad Hoc"], key="p3_otype")
        st.markdown("**Enter names to assign political orders (round-robin, max 1 difference)**")
        pol_names_raw = st.text_area("Names (one per line)",
                                     value="\n".join(st.session_state.pol_names),
                                     height=150, key="p3_names",
                                     placeholder="Alice\nBob\nCarol\n...")
        pol_names = [n.strip() for n in pol_names_raw.splitlines() if n.strip()]
        st.session_state.pol_names = pol_names
        total_pol = len(pol_df)
        if pol_names:
            base = total_pol // len(pol_names); extra = total_pol % len(pol_names)
            st.info(f"📊 **{total_pol}** orders ÷ **{len(pol_names)}** people = **{base}** each"
                    + (f", first **{extra}** person(s) get 1 extra" if extra else " (perfectly even)"))
        st.divider()
        st.markdown('<div class="step-header">Step 3 — Run political assignment</div>', unsafe_allow_html=True)
        can_run_pol = len(pol_names) > 0
        if not can_run_pol:
            st.info("ℹ️ Enter at least one name to enable assignment.")
        if st.button("🚀 Run Political Assignment", disabled=not can_run_pol, key="p3_run"):
            try:
                norm_pol = normalize_columns(pol_df, pol_order_type)
                norm_pol = norm_pol.reset_index(drop=True)
                assigned_pol = assign_round_robin(norm_pol, [pol_names], [100])
                out_pol = build_output(assigned_pol, pol_order_type)
                st.session_state.pol_processed = out_pol
                st.success(f"✅ Political assignment complete — {len(out_pol):,} orders assigned.")
            except Exception as e:
                st.error(f"Assignment failed: {e}"); st.exception(e)
        if st.session_state.pol_processed is not None:
            out_pol = st.session_state.pol_processed
            total_pol_out = len(out_pol)
            assigned_pol_n = int((out_pol["Names"] != "").sum())
            tim_pol = int(out_pol["TIM#"].nunique()) if "TIM#" in out_pol.columns else 0
            st.markdown("")
            mp1,mp2,mp3 = st.columns(3)
            for col,val,lbl in [(mp1,total_pol_out,"Total Political Orders"),(mp2,assigned_pol_n,"Assigned"),(mp3,tim_pol,"Unique TIM#")]:
                with col:
                    st.markdown(f'<div class="metric-card"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>', unsafe_allow_html=True)
            st.markdown("")
            pol_pivot = pd.DataFrame()
            if "Names" in out_pol.columns and "TIM#" in out_pol.columns:
                pol_pivot = (out_pol[out_pol["Names"]!=""].groupby("Names")["TIM#"].count()
                             .reset_index().rename(columns={"TIM#":"Total TIM# Count"})
                             .sort_values("Total TIM# Count",ascending=False).reset_index(drop=True))
            pt1,pt2,pt3 = st.tabs(["📄 Assigned Output","📊 Pivot","🔍 Raw Data"])
            with pt1: st.dataframe(out_pol, use_container_width=True, height=400)
            with pt2:
                if not pol_pivot.empty: st.dataframe(pol_pivot, use_container_width=True, height=400)
                else: st.info("No pivot data.")
            with pt3: st.dataframe(st.session_state.pol_raw_df, use_container_width=True, height=400)
            st.divider()
            st.markdown('<div class="step-header">Step 4 — Export</div>', unsafe_allow_html=True)
            ec1,ec2 = st.columns(2)
            with ec1:
                try:
                    xlsx_bytes = make_excel_bytes(out_pol, pol_pivot.copy() if not pol_pivot.empty else pd.DataFrame())
                    st.download_button("⬇️ Download Excel (.xlsx)", data=xlsx_bytes,
                                       file_name=f"political_orders_{date.today()}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True, key="p3_dl_xlsx")
                except Exception as e:
                    st.error(f"Excel generation failed: {e}"); st.exception(e)
            with ec2:
                st.download_button("⬇️ Download CSV", data=out_pol.to_csv(index=False).encode("utf-8"),
                                   file_name=f"political_orders_{date.today()}.csv", mime="text/csv",
                                   use_container_width=True, key="p3_dl_csv")
    else:
        st.markdown("""
        <div class="section-box" style="text-align:center;padding:2rem;">
            <div style="font-size:2rem;margin-bottom:0.4rem;">🏛️</div>
            <div style="color:#94a3b8;font-size:0.9rem;">
                Upload your political orders sheet and enter the team names.<br>
                Orders will be distributed equally (round-robin).
            </div>
            <hr style="border-color:#2d3148;margin:1rem auto;width:50%;">
            <div style="display:flex;justify-content:center;flex-wrap:wrap;gap:6px;">
                <span class="info-chip">✔ Handles small batches (10–15 orders/day)</span>
                <span class="info-chip">✔ True round-robin — max 1 order difference</span>
                <span class="info-chip">✔ Styled Excel + CSV export</span>
            </div>
        </div>""", unsafe_allow_html=True)
